import pandas as pd
import numpy as numpy
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
import tkinter as tk
from tkinter import messagebox
from tkinter import *
from openpyxl import *
from datetime import date

date11=date.today()
wb = load_workbook('diabetisedata.xlsx') 
sheet = wb.active 
df=pd.read_csv('C:/Users/Hp/Desktop/miniproject/diabetes.csv')
def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
	
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 30
	sheet.column_dimensions['C'].width = 30
	sheet.column_dimensions['D'].width = 30
	sheet.column_dimensions['E'].width = 30
	sheet.column_dimensions['F'].width = 30
	sheet.column_dimensions['G'].width = 30
	sheet.column_dimensions['H'].width = 30
	sheet.column_dimensions['I'].width = 30
	sheet.column_dimensions['J'].width = 30
	sheet.column_dimensions['k'].width=30
  
    # write given data to an excel spreadsheet 
    # at particular location 
	
	sheet.cell(row=1, column=1).value = "Date"
	sheet.cell(row=1, column=2).value = "Pregnancies"
	sheet.cell(row=1, column=3).value = "Glucose"
	sheet.cell(row=1, column=4).value = "BloodPressure"
	sheet.cell(row=1, column=5).value = "SkinThickness"
	sheet.cell(row=1, column=6).value = "Insulin"
	sheet.cell(row=1, column=7).value = "BMI"
	sheet.cell(row=1, column=8).value = "DiabetesPedigreeFunction"
	sheet.cell(row=1, column=9).value = "Age"
	sheet.cell(row=1, column=10).value = "Outcome"
	
	sheet.cell(row=1, column=11).value = "Name"

def validate(entries):
	datas=[]
	for i in entries:
		r = entries[i].get()
		if r ==' ':
			return False
		elif r=='':
			return False
	return True
	
excel()
print(df)
x=df.drop('Outcome',axis=1)

y=df['Outcome']
x_train=[]
y_train=[]
x_test=[]
y_test=[]
x_train,x_test,y_train,y_test=train_test_split(x,y,test_size=0.2,random_state=0)

lr=LogisticRegression()

lr.fit(x_train,y_train)

pred=lr.predict(x_test)

from sklearn import metrics

print("Percentage:-",metrics.accuracy_score(y_test,pred)*100,'%')



import tkinter as tk

fields = ('Pregnancies', 'Glucose', 'BloodPressure', 'SkinThickness', 'Insulin',' BMI','DiabetesPedigreeFunction','Age','Name')

def LogisticRegrr(entries):
	if validate(entries):
		info=" "
		
		
		data=[]
		for i in entries:
			r = entries[i].get()
			data.append(r)
	  
		ddd=pd.DataFrame({'Pregnancies':[data[0]],'Glucose':[data[1]],'BloodPressure':[data[2]],'SkinThickness':[data[3]],'Insulin':[data[4]],'BMI':[data[5]],'DiabetesPedigreeFunction':[data[6]],'Age':[data[7]]})
		pred=lr.predict(ddd)
		msg=func(pred)

		if pred==1:
			Info="Dear User You Have: "+data[0]+" Pregnancies"+"\n\t\t"+data[1]+" Glucose"+"\n\t\t"+data[2]+" BloodPreassure"+"\n\t\t"+data[3]+" Skinthickness"+"\n\t\t"+data[4]+" Insulin"+"\n\t\t"+data[5]+" BMI"+"\n\t\t"+data[6]+" DiabetesPedigreeFunction"+"\n\t\t"+data[7]+" Age"+"\n\n"+"\t\tFor This Data You Have :Diabetese"+"\n\tSo Please Take A Note Of This...."
			print(Info)
		else:
			Info="Dear User You Have: "+data[0]+" Pregnancies"+"\n\t\t"+data[1]+" Glucose"+"\n\t\t"+data[2]+" BloodPreassure"+"\n\t\t"+data[3]+" Skinthickness"+"\n\t\t"+data[4]+" Insulin"+"\n\t\t"+data[5]+" BMI"+"\n\t\t"+data[6]+" DiabetesPedigreeFunction"+"\n\t\t"+data[7]+" Age"+"\n\n"+"\t\tFor This Data You Have :No Diabetese"+"\n\tSo Please Take A Note Of This...."
			print(Info)
		
		
	  


		messagebox.showinfo("Analysis Result",Info)
		
		
		current_row = sheet.max_row 
		current_column = sheet.max_column 
		
		sheet.cell(row=current_row + 1, column=1).value = date11 
		sheet.cell(row=current_row + 1, column=2).value = data[0] 
		sheet.cell(row=current_row + 1, column=3).value = data[1] 
		sheet.cell(row=current_row + 1, column=4).value = data[2] 
		sheet.cell(row=current_row + 1, column=5).value = data[3] 
		sheet.cell(row=current_row + 1, column=6).value = data[4] 
		sheet.cell(row=current_row + 1, column=7).value = data[5] 
		sheet.cell(row=current_row + 1, column=8).value = data[6] 
		sheet.cell(row=current_row + 1, column=9).value = data[7] 
		sheet.cell(row=current_row + 1, column=10).value = func(pred)
		sheet.cell(row=current_row+1,column=11).value=data[8]
		wb.save('diabetisedata.xlsx') 
		print("Data INserted Successfully")
	else:
		messagebox.showinfo("Alert","All The Fiels Are Necessary")
def func(i):
    if(i==0):
        return 'Not Diabetic'
    else:
        return 'Diabetic'


def makeform(root, fields):
	entries = {}
	for field in fields:
		row = tk.Frame(root)
		lab = tk.Label(row, width=22,text=field+": ", anchor='w',bd=5)
		lab.config(font=('times',26,'italic'))
		ent = tk.Entry(row,bg='cyan',bd=5)
		ent.config(font=('times',26,'italic'))
		ent.insert(0, " ")
		row.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
		lab.pack(side=tk.LEFT)
		ent.pack(side=tk.RIGHT, expand=tk.YES, fill=tk.X)
		entries[field] = ent
	return entries

if __name__ == '__main__':
	root = tk.Tk()
	root.geometry("{}x{}+0+0".format(root.winfo_screenwidth(),root.winfo_screenheight()))
	var = StringVar()
	lab = tk.Label(text="Diabetes Analysis Model",height=5,width=50,bd=5,justify="center")

	lab.config(font=('times',26,'italic'))
	lab.pack()


	a=tk.Frame(root)
	a.pack()

	ents = makeform(a, fields)
		



		
	b1 = tk.Button(root, text='Predict data',bg='red',height=2,activebackground='green',width=10,bd=5,command=(lambda e=ents: LogisticRegrr(e)))
	b1.config(font=('times',26,'italic'))
	b1.pack(side=tk.LEFT)

	b3 = tk.Button(root, text='Quit',height=2,width=10,bg='red',activebackground='green',bd=5, command=root.quit)
	b3.config(font=('times',26,'italic'))
	b3.pack(side=tk.LEFT)
	root.mainloop()	