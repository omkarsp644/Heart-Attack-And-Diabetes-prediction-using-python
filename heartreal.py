import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import rcParams
from matplotlib.cm import rainbow
import warnings
warnings.filterwarnings('ignore')

from openpyxl import *
from datetime import date

from tkinter import messagebox

from sklearn.model_selection import train_test_split

from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from openpyxl import *
date11=date.today()
wb = load_workbook('heartdata.xlsx') 
sheet = wb.active 




def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
	
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 20
	sheet.column_dimensions['C'].width = 20
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 20
	sheet.column_dimensions['G'].width = 20
	sheet.column_dimensions['H'].width = 20
	sheet.column_dimensions['I'].width = 20
	sheet.column_dimensions['J'].width = 20
	sheet.column_dimensions['K'].width = 20
	sheet.column_dimensions['L'].width = 20
	sheet.column_dimensions['M'].width = 20
	sheet.column_dimensions['N'].width = 20
	sheet.column_dimensions['O'].width = 20
	sheet.column_dimensions['P'].width = 50
	sheet.column_dimensions['R'].width = 50
	
	
    # write given data to an excel spreadsheet 
    # at particular location 
	
	sheet.cell(row=1, column=1).value = "Date"
	sheet.cell(row=1, column=2).value = "age"
	sheet.cell(row=1, column=3).value = "sex"
	sheet.cell(row=1, column=4).value = "cp"
	sheet.cell(row=1, column=5).value = "trestbps"
	sheet.cell(row=1, column=6).value = "chol"
	sheet.cell(row=1, column=7).value = "fbs"
	sheet.cell(row=1, column=8).value = "restecg"
	sheet.cell(row=1, column=9).value = "thalach"
	sheet.cell(row=1, column=10).value = "exang"
	sheet.cell(row=1, column=11).value = "oldpeak"
	sheet.cell(row=1, column=12).value = "slope"
	sheet.cell(row=1, column=13).value = "ca"
	sheet.cell(row=1, column=14).value = "thal"
	sheet.cell(row=1, column=15).value="target"
	sheet.cell(row=1, column=16).value="Algorithm"
	sheet.cell(row=1, column=17).value="Name"
	

excel()
def validate(entries):
	datas=[]
	for i in entries:
		r = entries[i].get()
		if r ==' ':
			return False
		elif r=='':
			return False
	return True
	
from tkinter import *
import tkinter as tk
data = pd.read_csv('C:/Users/Hp/Desktop/miniproject/dataset.csv')

print(data)

y = data['target']
X = data.drop(['target'], axis = 1)
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.33, random_state = 0)

def insert(data,ans,algo):

	current_row = sheet.max_row 
	current_column = sheet.max_column 
	
	sheet.cell(row=current_row + 1, column=1).value = date11 
	sheet.cell(row=current_row + 1, column=2).value = data[0] 
	sheet.cell(row=current_row + 1, column=3).value = data[1] 
	sheet.cell(row=current_row + 1, column=4).value = data[2] 
	sheet.cell(row=current_row + 1, column=5).value = data[3] 
	sheet.cell(row=current_row + 1, column=6).value = data[4] 
	sheet.cell(row=current_row + 1, column=7).value = data[5] 
	sheet.cell(row=current_row + 1, column=8).value = data[6] 
	sheet.cell(row=current_row + 1, column=9).value = data[7] 
	sheet.cell(row=current_row + 1, column=10).value = data[8] 
	sheet.cell(row=current_row + 1, column=11).value = data[9] 
	sheet.cell(row=current_row + 1, column=12).value = data[10] 
	sheet.cell(row=current_row + 1, column=13).value = data[11] 
	sheet.cell(row=current_row + 1, column=14).value = data[12] 
	sheet.cell(row=current_row + 1, column=15).value = ans 
	sheet.cell(row=current_row + 1, column=16).value = algo
	sheet.cell(row=current_row + 1, column=17).value = entname.get()
	
	
	wb.save('heartdata.xlsx') 
	print("Data INserted Successfully")

def knn(entries):
	if validate(entries):
		info=" "
		datas=[]
		for i in entries:
			r = entries[i].get()
			datas.append(r)
	  
		dataset=pd.DataFrame({'age':[datas[0]],'sex':[datas[1]],'cp':[datas[2]],'trestbps':[datas[3]],'chol':[datas[4]],'fbs':[datas[5]],'restecg':[datas[6]],'thalach':[datas[7]],'exang':[datas[8]],'oldpeak':[datas[9]],'slope':[datas[10]],'ca':[datas[11]],'thal':[datas[12]]})
		
		print("\n\n**** KNeighborsClassifier ****")
		print("--------------------------------------------------------------------------------------")
		knn_scores = []
		for k in range(1,21):
			knn_classifier = KNeighborsClassifier(n_neighbors = k)
			knn_classifier.fit(X_train, y_train)
			knn_scores.append(knn_classifier.score(X_test, y_test))
		knn_scores.sort(reverse=True)
		
		print("The score for KNeighborsClassifier is {}% ".format(knn_scores[0]*100),"\n\n")
		msg="Analysed Result: "+func(knn_classifier.predict(dataset))
		messagebox.showinfo("Predicted:",msg)
		print("Analysed Result: ",knn_classifier.predict(dataset),"\n\n\n\n")
		insert(datas,func(knn_classifier.predict(dataset)),"KNeighbourClassifier")
		datas.clear()
		print("--------------------------------------------------------------------------------------\n\n\n\n")
	else:
		messagebox.showinfo("Alert","Enter All The Values")
	
	#####
	plt.plot([k for k in range(1, 21)], knn_scores, color = 'red')
	for i in range(1,21):
		plt.text(i, knn_scores[i-1], (i, knn_scores[i-1]))
	plt.xticks([i for i in range(1, 21)])
	plt.xlabel('Number of Neighbors (K)')
	plt.ylabel('Scores')
	plt.title('K Neighbors Classifier scores for different K values')
	plt.show()
	
	

def func(i):
	if(i==0):
		return 'Not Heart Attack'
	else:
		return 'Heart Attack'

def Dtc(entries):
	if validate(entries):

		datas=[]
		for i in entries:
			r = entries[i].get()
			datas.append(r)
	  
		print(datas)
		dataset=pd.DataFrame({'age':[datas[0]],'sex':[datas[1]],'cp':[datas[2]],'trestbps':[datas[3]],'chol':[datas[4]],'fbs':[datas[5]],'restecg':[datas[6]],'thalach':[datas[7]],'exang':[datas[8]],'oldpeak':[datas[9]],'slope':[datas[10]],'ca':[datas[11]],'thal':[datas[12]]})
		
		
		print("\n\n**** DECISION TREE CLASSIFIER ****")
		print("--------------------------------------------------------------------------------------")

		dt_scores = []
		for i in range(1, len(X.columns) + 1):
			dt_classifier = DecisionTreeClassifier(max_features = i, random_state = 0)
			dt_classifier.fit(X_train, y_train)
			dt_scores.append(dt_classifier.score(X_test, y_test))
		dt_scores.sort(reverse=True)
		print("The score for Decision Tree Classifier is {}%".format(dt_scores[0]*100),"\n\n\n")
		
		
		a=dt_classifier.predict(dataset)
		print("Analysed Result : ",a,"\n\n\n\n")
		msg="Analysed Result: "+func(a)
		messagebox.showinfo("Predicted:",msg)
		print("--------------------------------------------------------------------------------------\n\n\n\n")
		
		insert(datas,func(a),"DecisionTreeClassifier")
		datas.clear()
	else:
		messagebox.showinfo("Alert","All Fields Are Required..")
	
	######
	plt.plot([i for i in range(1, len(X.columns) + 1)], dt_scores, color = 'green')
	for i in range(1, len(X.columns) + 1):
		plt.text(i, dt_scores[i-1], (i, dt_scores[i-1]))
	plt.xticks([i for i in range(1, len(X.columns) + 1)])
	plt.xlabel('Max features')
	plt.ylabel('Scores')
	plt.title('Decision Tree Classifier scores for different number of maximum features')
	plt.show()
	
def rfc(entries):


	if validate(entries):
		datas=[]
		for i in entries:
			r = entries[i].get()
			datas.append(r)
	  
		dataset=pd.DataFrame({'age':[datas[0]],'sex':[datas[1]],'cp':[datas[2]],'trestbps':[datas[3]],'chol':[datas[4]],'fbs':[datas[5]],'restecg':[datas[6]],'thalach':[datas[7]],'exang':[datas[8]],'oldpeak':[datas[9]],'slope':[datas[10]],'ca':[datas[11]],'thal':[datas[12]]})
		
		print("\n\n**** RANDOM FOREST CLASSIFIER ****")
		print("--------------------------------------------------------------------------------------")

		rf_scores = []
		estimators = [10, 100, 200, 500, 1000]
		for i in estimators:	
			rf_classifier = RandomForestClassifier(n_estimators = i, random_state = 10)
			rf_classifier.fit(X_train, y_train)
			rf_scores.append(rf_classifier.score(X_test, y_test))
		rf_scores.sort(reverse=True)
		print("The score for Random Forest Classifier is {}% .".format(rf_scores[0]*100),"\n\n\n")
		
		
		a=rf_classifier.predict(dataset)
		msg="Analysed Result: "+func(rf_classifier.predict(dataset))
		messagebox.showinfo("Predicted:",msg)
		print("Analysed Result : ",a,"\n\n\n\n")
		print("--------------------------------------------------------------------------------------\n\n\n\n")
		
		insert(datas,func(a),"RandomForestClassifier")
		datas.clear()
	else:
		messagebox.showinfo("Alert","All Fields Are Required...")
	########	
	colors = rainbow(np.linspace(0, 1, len(estimators)))
	plt.bar([i for i in range(len(estimators))], rf_scores, color = colors, width = 0.8)
	for i in range(len(estimators)):
		plt.text(i, rf_scores[i], rf_scores[i])
	plt.xticks(ticks = [i for i in range(len(estimators))], labels = [str(estimator) for estimator in estimators])
	plt.xlabel('Number of estimators')
	plt.ylabel('Scores')
	plt.title('Random Forest Classifier scores for different number of estimators')
	plt.show()
	
def makeform(root, fields):
    entries = {}
    for field in fields:
        row = tk.Frame(root)
        lab = tk.Label(row, width=32,text=field+": ", anchor='w',bd=5)
        lab.config(font=('times',22,'italic'))
        ent = tk.Entry(row,bg='cyan',bd=5)
        ent.config(font=('times',26,'italic'))
        ent.insert(0, " ")
        row.pack(side=tk.TOP, 
                 fill=tk.X, 
                 padx=1, 
                 pady=1)
        lab.pack(side=tk.LEFT)
        ent.pack(side=tk.RIGHT, 
                 expand=tk.YES, 
                 fill=tk.X)
        entries[field] = ent
    return entries

if __name__ == '__main__':

	


	data1=data.drop(['target'],axis=1)
	
	fields=['age', 'sex', 'Chest Pain   (0-3)', 'Resting Blood Pressure (mmhg)', 'Cholestrol (mg/dl)', 'fasting blood sugar(if >120mg/dl then 1)', 'Resting ECG(0-2)', 'Max Heart Rate','Exercise Angina(0-1)', 'ST Depression', 'Peak Exe Slope', 'Major Vessels No', 'Thalassemia(3-6-7)']

	print(type(fields))


	root = tk.Tk()
	var = StringVar()
	root.geometry("{}x{}+0+0".format(root.winfo_screenwidth(),root.winfo_screenheight()))
	lab = tk.Label(text="Heart Disease Analysis Model",width=50,bd=5,justify="center")
	lab.config(font=('times',26,'italic'))
	lab.pack()
	a=tk.Frame(root)
	a.pack()

	
	row = tk.Frame(a)
	lab = tk.Label(row, width=32,text='Name'+": ", anchor='w',bd=5)
	lab.config(font=('times',22,'italic'))
	entname = tk.Entry(row,bg='cyan',bd=5)
	entname.config(font=('times',26,'italic'))
	entname.insert(0, " ")
	row.pack(side=tk.TOP,fill=tk.X, padx=1, pady=1)
	lab.pack(side=tk.LEFT)
	entname.pack(side=tk.RIGHT, expand=tk.YES, fill=tk.X)

	ents = makeform(a, fields)
	text=tk.Label(root,text="THEIR IS A TEXT:- WAIT AND WATCH",textvariable=var)
	text.config(font=('times',26,'italic'))
	text.pack(side=tk.LEFT)
		

	

		
	b1 = tk.Button(root, text='KNeighborsClassifier',bg='red',height=2,activebackground='green',bd=5,command=(lambda e=ents: knn(e)),anchor=CENTER)
	b1.config(font=('times',15,'italic'))
	b1.pack(side=tk.LEFT)
	b4 = tk.Button(root, text='Decision Tree Classifier',bg='yellow',height=2,activebackground='green',bd=5,command=(lambda e=ents: Dtc(e)),anchor=CENTER)
	b4.config(font=('times',15,'italic'))
	b4.pack(side=tk.LEFT)
	b5 = tk.Button(root, text='Random Forest Classifier',bg='white',height=2,activebackground='green',bd=5,command=(lambda e=ents: rfc(e)),anchor=CENTER)
	b5.config(font=('times',15,'italic'))
	b5.pack(side=tk.LEFT)

	b2 = tk.Button(root, text='Quit',height=2,bg='red',activebackground='green',bd=5, command=root.quit,anchor=CENTER)
	b2.config(font=('times',15,'italic'))
	b2.pack( side=tk.LEFT)
	
	root.mainloop()